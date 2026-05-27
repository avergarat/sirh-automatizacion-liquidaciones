# -*- coding: utf-8 -*-
import time
import datetime
from pathlib import Path
import threading
import subprocess

from openpyxl import load_workbook
from pywinauto.application import Application
from pywinauto.keyboard import send_keys
from pywinauto import Desktop, mouse   # <-- incluye mouse

import tkinter as tk
from tkinter import ttk, filedialog, messagebox


# =========================================================
# VARIABLES GLOBALES
# =========================================================
RUTA_EXCEL = None
HOJA = "FOLIOS"
CARPETA_SALIDA = None          # Path, se setea desde la GUI
FOLIOS = []                    # lista de (folio, numero, rut, corr_pago)
STOP_FLAG = False
PERIODO_AAAAMM = None          # AAAAmm elegido en GUI
ROOT = None                    # referencia al tk.Tk() para updates thread-safe

# Offsets para el campo FOLIO dentro de la ventana de consulta
OFFSET_FOLIO_X = 286
OFFSET_FOLIO_Y = 204

# Coordenadas ABSOLUTAS medidas para imprimir y OK
PRINT_ABS_X = 249    # botón imprimir en barra de Preview
PRINT_ABS_Y = 38
OK_ABS_X = 1083      # botón OK del cuadro de impresión
OK_ABS_Y = 467

# Coordenadas ABSOLUTAS para cerrar la página de la liquidación (botón cerrar)
CLOSE_ABS_X = 1894
CLOSE_ABS_Y = 7


# =========================================================
# CARGA DE FOLIOS DESDE EXCEL
# Lee: FOLIO, NUMERO, RUT, CORRELATIVO DE PAGO
# =========================================================
def cargar_folios(ruta_excel: str):
    wb = load_workbook(ruta_excel, data_only=True, read_only=True)
    ws = wb[HOJA]

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_folio = None
    col_numero = None
    col_rut = None
    col_corr_pago = None

    for i, name in enumerate(header, start=1):
        if not name:
            continue
        name_up = str(name).strip().upper()
        if name_up == "FOLIO":
            col_folio = i
        elif name_up == "NUMERO":
            col_numero = i
        elif name_up == "RUT":
            col_rut = i
        elif name_up == "CORRELATIVO DE PAGO":
            col_corr_pago = i

    if not col_folio:
        raise ValueError("No se encontró columna 'FOLIO' en la primera fila.")
    if not col_numero:
        raise ValueError("No se encontró columna 'NUMERO' en la primera fila.")
    if not col_rut:
        raise ValueError("No se encontró columna 'RUT' en la primera fila.")
    if not col_corr_pago:
        raise ValueError("No se encontró columna 'CORRELATIVO DE PAGO' en la primera fila.")

    datos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        folio_val = row[col_folio - 1]
        numero_val = row[col_numero - 1]
        rut_val = row[col_rut - 1]
        corr_val = row[col_corr_pago - 1]

        if folio_val is None or str(folio_val).strip() == "":
            continue

        folio = str(folio_val).strip()
        numero = str(numero_val).strip() if numero_val is not None else ""
        rut = str(rut_val).strip() if rut_val is not None else ""
        corr_pago = str(corr_val).strip() if corr_val is not None else ""

        datos.append((folio, numero, rut, corr_pago))

    if not datos:
        raise ValueError("No se encontraron filas con FOLIO en la hoja FOLIOS.")

    return datos


# =========================================================
# CONECTAR A LA VENTANA DE SIRH YA ABIERTA
# =========================================================
def conectar_sirh():
    try:
        app = Application(backend="win32").connect(
            title_re=r".*PAGO DE PLANILLAS ACCESORIAS.*Consulta de liquidaci[oó]n por folio.*",
            timeout=10,
        )
    except Exception as e:
        raise RuntimeError(
            "No encuentro la ventana de 'Consulta de liquidación por folio'. "
            f"Asegúrate de que esté abierta y al frente. Detalle: {e}"
        )

    win = app.top_window()
    win.wait("visible", timeout=10)
    win.set_focus()
    time.sleep(0.5)
    print("DEBUG ventana enganchada:", win.window_text())
    return win


# =========================================================
# ESCRIBIR FOLIO POR POSICIÓN
# =========================================================
def escribir_folio_por_posicion(ventana, folio: str):
    rect = ventana.rectangle()
    x = rect.left + OFFSET_FOLIO_X
    y = rect.top + OFFSET_FOLIO_Y

    ventana.set_focus()
    time.sleep(0.1)
    ventana.click_input(coords=(x, y))
    time.sleep(0.2)

    send_keys("^a{DEL}")
    time.sleep(0.1)
    send_keys(folio)
    time.sleep(0.1)


# =========================================================
# CLICK EN BOTÓN "LIMPIAR"
# =========================================================
def click_limpiar(ventana):
    try:
        ventana.set_focus()
        time.sleep(0.2)

        btn = ventana.child_window(control_id=138)
        btn = btn.wrapper_object()

        print("DEBUG Limpiar rect:", btn.rectangle())
        btn.click_input()
        time.sleep(0.5)
    except Exception as e:
        print("ADVERTENCIA: no pude pulsar el botón 'Limpiar' (id=138):", e)


# =========================================================
# GENERAR NOMBRE DE ARCHIVO (MANEJO DE REPETIDOS)
# Formato: AAAAmm_RUT_CORRELATIVO_DE_PAGO_A[_n].pdf
# =========================================================
def generar_nombre_pdf_unico(carpeta: Path, rut: str, corr_pago: str) -> Path:
    global PERIODO_AAAAMM

    if not PERIODO_AAAAMM:
        raise RuntimeError("PERIODO_AAAAMM no está definido.")

    rut_str = rut.strip()
    corr_str = corr_pago.strip()

    base = f"{PERIODO_AAAAMM}_{rut_str}_{corr_str}_A"
    contador = 0

    while True:
        nombre = f"{base}.pdf" if contador == 0 else f"{base}_{contador}.pdf"
        ruta = carpeta / nombre
        if not ruta.exists():
            return ruta
        contador += 1


# =========================================================
# PASO 1 + PASO 2: CLICK IMPRIMIR + CLICK OK
# (coordenadas absolutas)
# =========================================================
def ejecutar_pasos_imprimir_y_ok():
    # pequeña espera para que el preview termine de abrir
    time.sleep(3.0)

    # Paso 1: clic en icono de impresión
    print(f"PASO 1: clic imprimir en ({PRINT_ABS_X}, {PRINT_ABS_Y})")
    try:
        active = Desktop(backend="win32").get_active()
        print("DEBUG ventana activa antes de imprimir:", active.window_text())
    except Exception:
        pass
    mouse.click(button="left", coords=(PRINT_ABS_X, PRINT_ABS_Y))

    # Espera 0.5 segundo para que aparezca el cuadro de impresión
    time.sleep(0.5)

    # Paso 2: clic en botón OK del cuadro de impresión
    print(f"PASO 2: clic OK en ({OK_ABS_X}, {OK_ABS_Y})")
    mouse.click(button="left", coords=(OK_ABS_X, OK_ABS_Y))

    # pequeña espera para que se dispare el "Guardar como"
    time.sleep(0.5)


# =========================================================
# PROCESAR UN FOLIO COMPLETO
# =========================================================
def procesar_folio(ventana, folio: str, rut: str, corr_pago: str):
    global CARPETA_SALIDA

    print("Procesando FOLIO %s (RUT %s, CORR %s) ..." % (folio, rut, corr_pago))

    # 1) INGRESAR FOLIO
    escribir_folio_por_posicion(ventana, folio)

    # 2) ENTER
    send_keys("{ENTER}")
    time.sleep(0.8)

    # 3) BOTÓN "Copia Liqu." (control_id=137)
    try:
        boton_copia = ventana.child_window(control_id=137)
        w = boton_copia.wrapper_object()
        print("DEBUG rect Copia Liqu.:", w.rectangle())
        ventana.set_focus()
        time.sleep(0.1)
        w.click_input()
    except Exception as e:
        raise RuntimeError("No pude hacer clic en 'Copia Liqu.'. Detalle: %r" % e)

    # 👉 Aquí entras tú: clic icono impresión + clic OK en "Print".
    time.sleep(1)  # 🔹 retraso solicitado para mayor estabilidad
    ejecutar_pasos_imprimir_y_ok()

    # 4) ESPERAR EL DIÁLOGO "Guardar como"
    win_guardar = None
    for _ in range(80):  # hasta ~40s para que alcances a imprimir/OK
        if STOP_FLAG:          # permite abortar durante la espera
            return
        try:
            guardar_app = Application(backend="win32").connect(
                title_re=r".*(Guardar|Guardar como|Save As).*", timeout=0.5
            )
            win_guardar = guardar_app.top_window()
            win_guardar.wait("visible enabled ready", timeout=2)
            break
        except Exception:
            time.sleep(0.3)

    if STOP_FLAG:
        return

    if win_guardar is None:
        raise RuntimeError("No encontré el diálogo 'Guardar como' del PDF.")

    print("DEBUG diálogo Guardar como:", win_guardar.window_text())

    # 5) NOMBRE PDF ÚNICO: AAAAmm_RUT_CORR_PAGO_A[_n].pdf
    ruta_pdf = generar_nombre_pdf_unico(CARPETA_SALIDA, rut, corr_pago)
    destino = str(ruta_pdf)

    # 6) ESCRIBIR NOMBRE ARCHIVO
    try:
        cuadro_nombre = win_guardar.child_window(class_name="Edit")
        cuadro_nombre.set_edit_text(destino)
    except Exception:
        edits = [c for c in win_guardar.children() if c.friendly_class_name() == "Edit"]
        if not edits:
            raise RuntimeError("No encontré controles tipo 'Edit' en el diálogo Guardar.")
        cuadro_nombre = edits[0]
        cuadro_nombre.set_edit_text(destino)

    time.sleep(0.3)

    # 7) BOTÓN GUARDAR
    try:
        boton_guardar = win_guardar.child_window(
            title_re=r".*(Guardar|Save).*", class_name="Button"
        )
        boton_guardar.click_input()
    except Exception:
        posibles = ["&Guardar", "Guardar", "&Save", "Save"]
        click_ok = False
        for nombre in posibles:
            try:
                btn = win_guardar[nombre]
                btn.click()
                click_ok = True
                break
            except Exception:
                continue
        if not click_ok:
            raise RuntimeError("No pude hacer clic en 'Guardar' del diálogo.")

    time.sleep(0.5)

    # 👉 CLIC EXTRA PARA CERRAR LA PÁGINA DE LA LIQUIDACIÓN (según tus coordenadas)
    print(f"Cerrando página de liquidación en ({CLOSE_ABS_X}, {CLOSE_ABS_Y})")
    mouse.click(button="left", coords=(CLOSE_ABS_X, CLOSE_ABS_Y))
    time.sleep(0.5)

    # 8) CERRAR PREVIEW
    try:
        pdf_app = Application(backend="win32").connect(class_name="ThunderRT6FormDC", timeout=1)
        pdf_win = pdf_app.top_window()
        pdf_win.close()
    except Exception:
        pass

    # 9) LIMPIAR PARA SIGUIENTE FOLIO
    click_limpiar(ventana)


# =========================================================
# BUCLE PRINCIPAL CONTROLADO POR LA GUI
# =========================================================
def procesar_lote(desde_indice: int, textarea_status, boton_comenzar, boton_detener):
    global STOP_FLAG, FOLIOS, ROOT

    # Helper thread-safe: encola actualización de UI en el hilo principal
    def _log(msg):
        if ROOT:
            ROOT.after(0, lambda m=msg: (
                textarea_status.insert(tk.END, m + "\n"),
                textarea_status.see(tk.END),
            ))
        else:
            textarea_status.insert(tk.END, msg + "\n")
            textarea_status.see(tk.END)

    def _fin():
        boton_comenzar.config(state="normal")
        boton_detener.config(state="disabled")

    try:
        ventana = conectar_sirh()
    except Exception as e:
        _log(f"❌ Error conectando a SIRH: {e}")
        if ROOT:
            ROOT.after(0, _fin)
        else:
            _fin()
        return

    for i in range(desde_indice, len(FOLIOS)):
        if STOP_FLAG:
            _log("⏹ Proceso detenido por el usuario.")
            break

        folio, _, rut, corr_pago = FOLIOS[i]
        try:
            _log(f"➡ Fila {i+1}: FOLIO {folio} | RUT {rut} | CORR {corr_pago}")
            procesar_folio(ventana, folio, rut, corr_pago)
            if not STOP_FLAG:
                _log(f"✅ Terminado FOLIO {folio}")
        except Exception as e:
            _log(f"❌ Error en FOLIO {folio}: {e}")
            try:
                ventana.set_focus()
                click_limpiar(ventana)
            except Exception:
                pass
            time.sleep(1)

    if ROOT:
        ROOT.after(0, _fin)
    else:
        _fin()


# =========================================================
# GUI TKINTER
# =========================================================
def crear_gui():
    global RUTA_EXCEL, CARPETA_SALIDA, FOLIOS, STOP_FLAG, PERIODO_AAAAMM, ROOT

    COLOR_FONDO = "#0B1F33"      # azul oscuro (EVIDANT-ish)
    COLOR_PRIMARIO = "#00A6A6"   # turquesa
    COLOR_SECUNDARIO = "#F2F2F2" # gris claro

    root = tk.Tk()
    ROOT = root   # permite thread-safe updates desde procesar_lote
    root.title("EVIDANT - Automatización Liquidaciones")
    root.configure(bg=COLOR_FONDO)
    root.minsize(780, 520)

    # HEADER
    header = tk.Frame(root, bg=COLOR_FONDO)
    header.pack(fill="x", pady=10)
    lbl_title = tk.Label(
        header,
        text="EVIDANT · Descarga automática de liquidaciones",
        bg=COLOR_FONDO,
        fg="white",
        font=("Segoe UI", 14, "bold"),
    )
    lbl_title.pack()

    # CONFIG
    frame_cfg = tk.Frame(root, bg=COLOR_FONDO)
    frame_cfg.pack(fill="x", padx=20, pady=10)

    # Excel
    tk.Label(frame_cfg, text="Archivo Excel de folios:", bg=COLOR_FONDO, fg="white").grid(
        row=0, column=0, sticky="w"
    )
    entry_excel = tk.Entry(frame_cfg, width=60)
    entry_excel.grid(row=0, column=1, padx=5, pady=3, sticky="w")

    # Carpeta salida
    tk.Label(frame_cfg, text="Carpeta salida PDF:", bg=COLOR_FONDO, fg="white").grid(
        row=1, column=0, sticky="w"
    )
    entry_carpeta = tk.Entry(frame_cfg, width=60)
    entry_carpeta.grid(row=1, column=1, padx=5, pady=3, sticky="w")

    # Periodo AAAAmm
    tk.Label(frame_cfg, text="Periodo (AAAAmm):", bg=COLOR_FONDO, fg="white").grid(
        row=2, column=0, sticky="w"
    )
    entry_periodo = tk.Entry(frame_cfg, width=10)
    entry_periodo.grid(row=2, column=1, padx=5, pady=3, sticky="w")

    # valor por defecto: año/mes actual
    hoy = datetime.date.today()
    entry_periodo.insert(0, f"{hoy.year}{hoy.month:02d}")

    # FOLIO INICIO
    tk.Label(frame_cfg, text="Comenzar desde:", bg=COLOR_FONDO, fg="white").grid(
        row=3, column=0, sticky="w"
    )
    combo_folio_inicio = ttk.Combobox(frame_cfg, width=55, state="readonly")
    combo_folio_inicio.grid(row=3, column=1, padx=5, pady=3, sticky="w")

    # STATUS
    frame_status = tk.Frame(root, bg=COLOR_FONDO)
    frame_status.pack(fill="both", expand=True, padx=20, pady=10)
    textarea_status = tk.Text(frame_status, bg=COLOR_SECUNDARIO, fg="black", height=15)
    textarea_status.pack(fill="both", expand=True)

    # HANDLERS
    def seleccionar_excel():
        nonlocal entry_excel, combo_folio_inicio
        global RUTA_EXCEL, FOLIOS

        path = filedialog.askopenfilename(
            title="Seleccionar Excel de folios",
            filetypes=[("Archivos Excel", "*.xlsx;*.xlsm"), ("Todos", "*.*")],
        )
        if not path:
            return

        entry_excel.delete(0, tk.END)
        entry_excel.insert(0, path)

        try:
            folios = cargar_folios(path)
        except Exception as e:
            messagebox.showerror("Error al leer Excel", str(e))
            return

        FOLIOS = folios
        valores = [
            f"{idx+1:04d} - FOLIO {f} · RUT {rut} · PAGO {corr}"
            for idx, (f, n, rut, corr) in enumerate(FOLIOS)
        ]
        combo_folio_inicio["values"] = valores
        if valores:
            combo_folio_inicio.current(0)

        textarea_status.insert(tk.END, f"✔ Cargados {len(FOLIOS)} registros desde el Excel.\n")
        textarea_status.see(tk.END)
        RUTA_EXCEL = path

    def seleccionar_carpeta():
        nonlocal entry_carpeta
        global CARPETA_SALIDA, RUTA_EXCEL

        inicial = entry_carpeta.get().strip() or (
            str(Path(RUTA_EXCEL).parent) if RUTA_EXCEL else ""
        )
        # Escapar comillas simples para PowerShell
        inicial_ps = inicial.replace("'", "''")

        resultado = [None]

        def _picker():
            """Hilo: abre FolderBrowserDialog nativo via PowerShell (funciona en modo Admin)."""
            ps = (
                "Add-Type -AssemblyName System.Windows.Forms; "
                "$f = New-Object System.Windows.Forms.FolderBrowserDialog; "
                "$f.Description = 'Seleccionar carpeta de salida para PDFs'; "
                f"$f.SelectedPath = '{inicial_ps}'; "
                "$f.ShowNewFolderButton = $true; "
                "if ($f.ShowDialog() -eq 'OK') { Write-Output $f.SelectedPath }"
            )
            try:
                r = subprocess.run(
                    ["powershell", "-NoProfile", "-Command", ps],
                    capture_output=True, text=True, timeout=120,
                )
                resultado[0] = r.stdout.strip()
            except Exception:
                resultado[0] = ""
            root.after(0, _aplicar)

        def _aplicar():
            """Hilo principal: aplica la carpeta elegida."""
            global CARPETA_SALIDA
            if not resultado[0]:
                return
            carpeta = Path(resultado[0])
            try:
                carpeta.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo crear la carpeta:\n{e}", parent=root)
                return
            CARPETA_SALIDA = carpeta
            entry_carpeta.delete(0, tk.END)
            entry_carpeta.insert(0, str(carpeta))
            textarea_status.insert(tk.END, f"✔ Carpeta de salida: {carpeta}\n")
            textarea_status.see(tk.END)

        threading.Thread(target=_picker, daemon=True).start()

    # BOTONES BUSCAR/DEFINIR
    btn_excel = tk.Button(
        frame_cfg,
        text="Buscar...",
        command=seleccionar_excel,
        bg=COLOR_PRIMARIO,
        fg="white",
        relief="flat",
    )
    btn_excel.grid(row=0, column=2, padx=5, pady=3)

    btn_carpeta = tk.Button(
        frame_cfg,
        text="Seleccionar",
        command=seleccionar_carpeta,
        bg=COLOR_PRIMARIO,
        fg="white",
        relief="flat",
    )
    btn_carpeta.grid(row=1, column=2, padx=5, pady=3)

    # BOTONES CONTROL
    frame_btn = tk.Frame(root, bg=COLOR_FONDO)
    frame_btn.pack(fill="x", padx=20, pady=10)

    btn_comenzar = tk.Button(
        frame_btn,
        text="Comenzar",
        width=12,
        bg=COLOR_PRIMARIO,
        fg="white",
        relief="flat",
    )
    btn_detener = tk.Button(
        frame_btn,
        text="Detener",
        width=12,
        bg="#CC3333",
        fg="white",
        relief="flat",
        state="disabled",
    )
    btn_comenzar.pack(side="left", padx=5)
    btn_detener.pack(side="left", padx=5)

    def iniciar_proceso():
        global STOP_FLAG, CARPETA_SALIDA, FOLIOS, PERIODO_AAAAMM

        if not FOLIOS:
            messagebox.showwarning("Faltan datos", "Primero selecciona el Excel de folios.")
            return
        if CARPETA_SALIDA is None:
            messagebox.showwarning("Faltan datos", "Primero define la carpeta de salida.")
            return

        periodo = entry_periodo.get().strip()
        if len(periodo) != 6 or not periodo.isdigit():
            messagebox.showwarning(
                "Periodo inválido",
                "El periodo debe tener el formato AAAAmm (6 dígitos)."
            )
            return
        PERIODO_AAAAMM = periodo

        idx = combo_folio_inicio.current()
        if idx < 0:
            idx = 0

        STOP_FLAG = False
        btn_comenzar.config(state="disabled")
        btn_detener.config(state="normal")

        hilo = threading.Thread(
            target=procesar_lote,
            args=(idx, textarea_status, btn_comenzar, btn_detener),
            daemon=True,
        )
        hilo.start()

    def detener_proceso():
        global STOP_FLAG
        STOP_FLAG = True

    btn_comenzar.config(command=iniciar_proceso)
    btn_detener.config(command=detener_proceso)

    root.mainloop()


# =========================================================
# ENTRY POINT
# =========================================================
if __name__ == "__main__":
    crear_gui()
